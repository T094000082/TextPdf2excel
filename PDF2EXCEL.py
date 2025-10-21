from datetime import datetime
import pdfplumber
import openpyxl
import re

# Step 1: 擷取 PDF 文字
with pdfplumber.open(r"D:\LEO_Coding_Work\Python\Ocr_Collection\PDF2EXCEL\91E1-20240122001.pdf") as pdf:
    text = ""
    for page in pdf.pages:
        text += page.extract_text()

# 先輸出文字內容，方便除錯
print("=== PDF 文字內容 ===")
print(text)
print("===================\n")

# Step 2: 擷取金額
# 擷取高鐵票價 - 尋找 "票價金額(Fare,$):" 或 "票價金額(Fare,$):" 後面的數字
highspeed_matches = re.findall(r"票價金額\(Fare,\$\):\s*(\d+)", text)

# 如果找不到，嘗試其他可能的格式
if not highspeed_matches:
    highspeed_matches = re.findall(r"NT\$(\d+)", text)

# 擷取其他交通方式
other_transport = re.findall(r"(計程車|公務車|捷運).*?金額.*?(\d{1,3}(?:,\d{3})*)", text)

# 合併結果
matches = []
for price in highspeed_matches:
    matches.append(("高鐵", price))
matches.extend(other_transport)

#釋放檔案避免佔用
pdf.close()

# Step 3: 開啟 Excel 模板
wb = openpyxl.load_workbook("出差旅費報支單-空白.xlsx")
ws = wb.active

# Step 4: 寫入對應欄位
highspeed_total = 0
taxi_total = 0

for item, amount in matches:
    amount = float(amount.replace(",", ""))
    if "高鐵" in item:
        highspeed_total += amount
    elif "計程車" in item:
        taxi_total += amount

# 在Console輸出結果
print(f"高鐵總金額: {highspeed_total}")
print(f"計程車總金額: {taxi_total}")    
print(f"總金額: {highspeed_total + taxi_total}")

# 寫入 Excel（根據您的 Excel 欄位結構調整）
if highspeed_total > 0:
    ws["F7"].value = highspeed_total  # 高鐵金額
if taxi_total > 0:
    ws["J7"].value = taxi_total  # 計程車金額

# Step 5: 儲存結果 使用目前時間 YYYYMMdd-HHmm作為檔名
current_time = datetime.now().strftime("%Y%m%d-%H%M")
output_file = f"分析結果_{current_time}.xlsx"

wb.save(output_file)
